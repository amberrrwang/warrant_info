from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import openpyxl, os, re, time
import requests  # ← 新增：用來打 Yuanta API

# ======= 設定 =======
wid_list = [
    "03111U", "03162U", "03485U", "03616U", "03662U",
    "03281U", "03864U", "05831P", "063866", "065413", "071599",
    "07879P", "079683", "085398", "08700P", "08769P", "08992P",
    "71280U", "71286U", "71289U", "71344U", "71974U"
]

BASIC_LABELS = [
    "上市日期","最後交易日","到期日期","發行型態","最新發行張數",
    "流通在外張數/比例","最新履約價","最新行使比例",
    "買價隱波","賣價隱波","Delta","Theta",
    "剩餘天數","價內外程度","實質槓桿","買賣價差比"
]

# 只保留「標的股價」，不再有「標的現價」
HEADER_ORDER = [
    "WID","狀態","成交價","買價","賣價",
    "標的名稱","標的股價","標的代碼",
    *BASIC_LABELS, "抓取時間","來源網址"
]

# ======= 啟動 Driver =======
def launch_driver(headless=False):
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

# ======= 抓資料輔助 =======
def text_or_blank(driver, by, sel):
    try:
        return driver.find_element(by, sel).text.strip()
    except NoSuchElementException:
        return ""

def find_basic_value_by_label(driver, label_text):
    xps = [
        f"//*[normalize-space(text())='{label_text}']/following-sibling::*[1]",
        f"//div[.//*[normalize-space(text())='{label_text}']]/*[normalize-space(text())='{label_text}']/following-sibling::*[1]",
        f"//li[.//*[normalize-space(text())='{label_text}']]//*[normalize-space(text())='{label_text}']/following::*[1]",
    ]
    for xp in xps:
        try:
            txt = driver.find_element(By.XPATH, xp).text.strip()
            if txt:
                return txt
        except NoSuchElementException:
            continue
    return ""

def get_target_name_code(driver):
    """抓標的名稱/代碼（不抓價）。"""
    name, code = "", ""

    # 名稱
    for xp in ["//*[contains(@ng-bind, 'TAR_NAME') or contains(@ng-bind, 'FLD_TAR_NAME')]"]:
        els = driver.find_elements(By.XPATH, xp)
        if els and els[0].text.strip():
            name = els[0].text.strip()
            break

    # 代碼
    for xp in ["//*[contains(@ng-bind, 'TAR_CODE') or contains(@ng-bind, 'FLD_TAR_CODE')]"]:
        els = driver.find_elements(By.XPATH, xp)
        if els and els[0].text.strip():
            code = re.sub(r"\D", "", els[0].text.strip())
            break

    # 備援：從含「標的」的文字解析
    if not (name and code):
        try:
            block = driver.find_element(By.XPATH, "//*[contains(normalize-space(.), '標的')]").text.strip()
            if not name:
                m_name = re.search(r"標的[:：]\s*([^\s／/｜|()（）]+)", block)
                name = m_name.group(1) if m_name else name
            if not code:
                m_code = re.search(r"\((\d{4})\)", block) or re.search(r"[^\d](\d{4})(?:\D|$)", block)
                code = m_code.group(1) if m_code else code
        except NoSuchElementException:
            pass

    return name, code

# ======= NEW：從 Yuanta API 取「標的股價＝賣一(ask1)」 =======
def _to_num(x):
    if x is None: 
        return None
    try:
        return float(str(x).replace(",", "").strip())
    except:
        return None

def get_udly_mid_from_api(udly_code: str, timeout=8):
    """五檔 API：回傳 (bid1+ask1)/2 中間價；若缺邊就回現有那邊；全失敗回 None。"""
    if not udly_code:
        return None
    url = f"https://www.warrantwin.com.tw/eyuanta/ws/Quote.ashx?type=mem_ta5&symbol={udly_code}"
    try:
        r = requests.get(url, timeout=timeout); r.raise_for_status()
        items = r.json().get("items", {})
        bid1 = items.get("101") if isinstance(items, dict) else None
        ask1 = items.get("102") if isinstance(items, dict) else None
        bid1, ask1 = _to_num(bid1), _to_num(ask1)
        if bid1 is not None and ask1 is not None:
            return (bid1 + ask1) / 2
        return ask1 if ask1 is not None else bid1
    except Exception:
        return None

def parse_conver_rate(v):
    """把『最新行使比例』欄位轉成 float；支援 '0.0050'、'0.5%'、'1/200' 之類字串。"""
    if v is None: 
        return None
    s = str(v).strip()
    # 分數
    if "/" in s:
        try:
            a, b = s.split("/", 1)
            return float(a) / float(b)
        except Exception:
            pass
    # 百分比
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", "")) / 100.0
        except Exception:
            pass
    # 一般數字
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None

def yuanta_calc_price(symbol: str, udly_price: float, conver_rate: float, war_type: int = 2, timeout=8):
    """
    直呼元大『試算』API，回傳 (PriceTheory, 全部原始 json)。
    war_type: 2=買權（網站上大多如此），1=賣權（若需要可改）。
    """
    url = "https://www.warrantwin.com.tw/eyuanta/ws/Quote.ashx"
    params = {
        "type": "calc",
        "symbol": symbol,
        "war_type": war_type,
        "conver_rate": f"{conver_rate:.6f}",
        "udly_price": f"{udly_price:.4f}",
    }
    r = requests.get(url, params=params, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    return data.get("PriceTheory"), data

# （可留作備援）從 DOM 五檔表抓第一列賣價
def get_target_best_ask_from_dom(driver):
    try:
        WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(normalize-space(.), '標的五檔報價')]"))
        )
        td = driver.find_element(
            By.XPATH, "//*[contains(normalize-space(.), '標的五檔報價')]/following::table[1]//tr[1]/td[3]"
        )
        return td.text.strip().replace(",", "")
    except Exception:
        return ""

def ensure_all_keys(row: dict) -> dict:
    for k in HEADER_ORDER:
        row.setdefault(k, "")
    return row

# ======= 抓單筆 =======
def scrape_one_wid(driver, wid):
    url = f"https://www.warrantwin.com.tw/eyuanta/Warrant/Info.aspx?WID={wid}"
    driver.get(url)

    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(@ng-bind, 'WAR_BUY_PRICE')]"))
        )
    except TimeoutException:
        return ensure_all_keys({
            "WID": wid, "狀態": "Timeout", "來源網址": url,
            "抓取時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    # 三價
    deal = text_or_blank(driver, By.XPATH, "//*[contains(@ng-bind, 'WAR_DEAL_PRICE')]")
    buy  = text_or_blank(driver, By.XPATH, "//*[contains(@ng-bind, 'WAR_BUY_PRICE')]")
    sell = text_or_blank(driver, By.XPATH, "//*[contains(@ng-bind, 'WAR_SELL_PRICE')]")

    # 標的名稱/代碼
    tgt_name, tgt_code = get_target_name_code(driver)

    # 標的股價：先用中間價（你也可改成成交價）
    udly_mid = get_udly_mid_from_api(tgt_code)

    row = {
        "WID": wid,
        "狀態": "OK",
        "成交價": deal,
        "買價": buy,
        "賣價": sell,
        "標的名稱": tgt_name,
        "標的股價": udly_mid,   # 供參考
        "標的代碼": tgt_code,
        "來源網址": url,
        "抓取時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # 基本欄
    for label in BASIC_LABELS:
        row[label] = find_basic_value_by_label(driver, label)

    # 解析行使比例（calc API 需要）
    conver_rate = parse_conver_rate(row.get("最新行使比例"))

    # 呼叫 calc API：用中間價（或你想要的 S）來求官網理論價
    price_theory = None
    if udly_mid is not None and conver_rate is not None:
        try:
            price_theory, _raw = yuanta_calc_price(
                symbol=wid,
                udly_price=float(udly_mid),
                conver_rate=float(conver_rate),
                war_type=2  # 買權
            )
        except Exception as e:
            print("⚠️ yuanta_calc_price error:", e)

    row["理論價(PriceTheory)"] = price_theory  # 這就是官網顯示的 1.3 那個數

    return ensure_all_keys(row)

# ======= 寫 Excel + 試算 =======
def save_rows_to_excel(rows, filename="yuanta_warrants.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "元大權證"

    # 寫表頭 + 資料
    ws.append(HEADER_ORDER)
    for r in rows:
        ws.append([r.get(k, "") for k in HEADER_ORDER])

    # 建立試算頁
    calc = wb.create_sheet("試算")

    # 標籤
    calc["A1"] = "WID"
    calc["A2"] = "標的股價"
    calc["A3"] = "買價隱波（％）"
    calc["A4"] = "評價日"
    calc["A6"] = "無風險利率 r（年化）"
    calc["F1"] = "（以下自動帶入）"
    calc["F2"] = "履約價 K"
    calc["F3"] = "剩餘天數"
    calc["F4"] = "行使比例（數值）"

    # 預設把 B1 設成第一筆的 WID（可用下拉切換）
    first_wid = rows[0].get("WID", "") if rows else ""
    calc["B1"] = first_wid

    # ---- 以 B1 的 WID 動態查找「元大權證」對應列 ----
    # 欄位位置（依你的 HEADER_ORDER）
    # A:WID, G:標的股價, O:最新履約價, P:最新行使比例, Q:買價隱波, U:剩餘天數
    sheet_name = "'元大權證'"
    calc["B2"] = f"=INDEX({sheet_name}!G:G, MATCH(B1, {sheet_name}!A:A, 0))"
    # 轉百分比字串為小數：SUBSTITUTE 去掉 %，VALUE 轉數字，再 /100
    calc["B3"] = (
    f"=IF(ISNUMBER(INDEX({sheet_name}!Q:Q, MATCH(B1, {sheet_name}!A:A, 0))),"
    f" INDEX({sheet_name}!Q:Q, MATCH(B1, {sheet_name}!A:A, 0)),"
    f" IF(RIGHT(INDEX({sheet_name}!Q:Q, MATCH(B1, {sheet_name}!A:A, 0)))=\"%\","
    f"     VALUE(SUBSTITUTE(INDEX({sheet_name}!Q:Q, MATCH(B1, {sheet_name}!A:A, 0)),\"%\",\"\"))/100,"
    f"     VALUE(INDEX({sheet_name}!Q:Q, MATCH(B1, {sheet_name}!A:A, 0)))"
    f" ))"
    )
    calc["B4"] = "=TODAY()"
    calc["B6"] = 0.01  # 你可自行改

    calc["G2"] = f"=INDEX({sheet_name}!O:O, MATCH(B1, {sheet_name}!A:A, 0))"
    calc["G3"] = f"=INDEX({sheet_name}!U:U, MATCH(B1, {sheet_name}!A:A, 0))"
    calc["G4"] = f"=INDEX({sheet_name}!P:P, MATCH(B1, {sheet_name}!A:A, 0))"

    # ---- Black–Scholes（買權）公式（乘上行使比例）----
    bs_formula = (
    "= ( B2*NORMSDIST((LN(B2/G2)+(B6+B3^2/2)*(G3/365))/(B3*SQRT(G3/365)))"
    " - G2*EXP(-B6*(G3/365))*NORMSDIST((LN(B2/G2)+(B6-B3^2/2)*(G3/365))/(B3*SQRT(G3/365))) ) * G4"
    )
    calc["C10"] = bs_formula
    calc["B10"] = "成交價："

    # 粗體 & 欄寬
    for cell in ["A1","A2","A3","A4","A6","F2","F3","F4","B10"]:
        calc[cell].font = openpyxl.styles.Font(bold=True)
    for col, width in [("A",16),("B",18),("C",28),("F",22),("G",18)]:
        calc.column_dimensions[col].width = width

    # WID 下拉資料驗證（從「元大權證」A2:A?）
    last_row = ws.max_row
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=f"={sheet_name}!$A$2:$A${last_row}", allow_blank=False)
    calc.add_data_validation(dv)
    dv.add(calc["B1"])

    # 存檔
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    out_path = os.path.join(desktop, filename)
    wb.save(out_path)
    print(f"✅ 已寫入 Excel：{out_path}")

# ======= 主流程 =======
def scrape_with_retry(driver, wid, retries=2, pause=1.2):
    for i in range(retries + 1):
        row = scrape_one_wid(driver, wid)
        if row.get("狀態") == "OK":
            return row
        time.sleep(pause)
    return row 
     
def main():
    driver = launch_driver(headless=False)
    rows = []
    try:
        for wid in wid_list:
            print(f"🔎 抓取 {wid} 中...")
            row = scrape_with_retry(driver, wid, retries=2, pause=1.2)
            print(
                f"→ 狀態:{row.get('狀態')} 成交:{row.get('成交價','')} 買:{row.get('買價','')} 賣:{row.get('賣價','')} | "
                f"標的代碼:{row.get('標的代碼','')} 標的股價(賣一):{row.get('標的股價','')}"
            )
            rows.append(row)
            time.sleep(1.0)  # 每筆之間也放慢
    finally:
        driver.quit()

    if rows:
        save_rows_to_excel(rows)
    else:
        print("⚠️ 沒有資料可寫入")

if __name__ == "__main__":
    main()


